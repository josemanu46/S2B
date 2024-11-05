import os.path
import time
import sys
from tkinter import messagebox
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
import openpyxl
import threading
import tkinter as tk
from main import get_result
import re, zipfile
from tkinter import *
import customtkinter
from tkinter import messagebox

class MainWindow:
    
    def __init__(self):
        self.createWindow()
        self.current_dir = str(os.path.dirname(os.path.realpath(__file__)))

    def createWindow(self):

        def on_closing():
            if messagebox.askyesno("S2B","Are you sure you want to go out?"):
                self.root.destroy()
                print("Thank you for using our software! Have a great day!")
                sys.exit()

        def create_excel_template():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            ws.cell(row=1, column=1, value="Contract")
            ws.cell(row=1, column=2, value="DU ID")
            ws.cell(row=1, column=3, value="Spart")
            ws.cell(row=1, column=4, value="QTY")

            self.download_folder = os.path.expanduser("~/Downloads")
            if not os.path.exists(self.download_folder):
                os.makedirs(self.download_folder)

            self.save_file_path = os.path.join(self.download_folder, "Plantilla_ML.xlsx")
            wb.save(self.save_file_path)

            return self.save_file_path


        def download_template():
            create_excel_template()
            time.sleep(2)
            print('waiting template 5s')
            try:
                os.startfile(self.save_file_path)
            except OSError as e:
                messagebox.showerror("Error", f"Could not open the folder file:\n{str(e)}")

        #window = tk.Tk()  # create Tk window like you do with the Tk window
        self.root = tk.Tk()

        self.root.configure(bg="#000129")
        self.root.title("S2B")
        self.root.protocol("WM_DELETE_WINDOW",on_closing)
        #window.wm_iconphoto(False, tkinter.PhotoImage(file='icon.png'))
        self.root.state('normal')

        windowWidth = 500
        windowHeight = 400
        positionRight = int((self.root.winfo_screenwidth() / 2) - ((windowWidth + 12) / 2))
        positionDown = int((self.root.winfo_screenheight() / 2) - ((windowHeight + 50) / 2))
        self.root.geometry('{}x{}+{}+{}'.format(windowWidth, windowHeight, positionRight, positionDown))
        #self.root.minsize(500,300)

        self.file_panel = tk.Frame(master=self.root)
        self.file_panel.configure(bg="#000129")
        self.file_panel.pack(padx=10, pady=(20, 10), fill="both", expand=False)
        self.file_panel.rowconfigure(0, weight=1)
        self.file_panel.rowconfigure(1, weight=1)
        self.file_panel.columnconfigure(0, weight=1)
        self.file_panel.columnconfigure(1, weight=1)

        #file_path_input_ml = tk.StringVar(value= "SELECT THE ML FILE")
        #self.entry_input_ml = tk.Entry(master=file_panel,
        #                        width=300,
        #                        textvariable=file_path_input_ml,
        #                        justify='center')
        #self.entry_input_ml.grid(row=0, column=0, pady=20, padx=20, ipadx=150, sticky="ew")
        #self.entry_input_ml.bind("<1>", lambda name: self.openPath(file_path_input_ml, "ml"))



        #self.select_files_button.pack(pady=(5, 20), padx=10)
        #self.select_files_button.config(bg="#1A2F43",highlightthickness=4,borderwidth=5,highlightbackground='white')
        self.select_files_button_Material_list = tk.Button(self.file_panel, text="Lista de Material", font=("San Francisco", 10),fg="#1A2F43",command=self.select_files_Material_list)
        self.select_files_button_Material_list.grid(row=1, column=0, pady=20, padx=20, ipadx=150, sticky="ew")   

        #self.entry_input_ml.bind("<1>", lambda name: self.select_files)

        self.select_files_button_restOrderBpartinfo = tk.Button(self.file_panel, text="Rest Order Bpart archivo", font=("San Francisco", 10),fg="#1A2F43",command=self.select_files_restOrderBpartinfo)
        self.select_files_button_restOrderBpartinfo.grid(row=2, column=0, pady=20, padx=20, ipadx=150, sticky="ew")

        #self.file_path_restOrderBpartinfo = tk.StringVar(value= "SELECT REST ORDER BPART FILE")
        #self.entry_restOrderBpartinfo = tk.Entry(master=self.file_panel,
        #                        width=300,
        #                        textvariable=self.file_path_restOrderBpartinfo,
        #                        justify='center')
        #self.entry_restOrderBpartinfo.grid(row=1, column=0, pady=20, padx=20, ipadx=150, sticky="ew")
        #self.entry_restOrderBpartinfo.bind("<1>", lambda name: self.openPath(self.file_path_restOrderBpartinfo, "restOrderBpartinfo"))

        self.select_files_button_remaining = tk.Button(self.file_panel, text="Remaining archivo", font=("San Francisco", 10),fg="#1A2F43",command=self.select_files_Remaining)
        self.select_files_button_remaining.grid(row=3, column=0, pady=20, padx=20, ipadx=150, sticky="ew")



        #self.file_path_remaining = tk.StringVar(value= "SELECT REMAINING FILE")
        #self.entry_remaining = tk.Entry(master=self.file_panel,
        #                        width=300,
        #                        textvariable=self.file_path_remaining,
        #                        justify='center')
        #self.entry_remaining.grid(row=2, column=0, pady=20, padx=20, ipadx=150, sticky='ew')
        #self.entry_remaining.bind("<1>", lambda name: self.openPath(self.file_path_remaining, "remaining"))

        # Botón para seleccionar archivos
        self.button_donwload = tk.Button(self.file_panel, text="Donwload Template", font=("San Francisco", 10),fg="#1A2F43",command=download_template)
        self.button_donwload.grid(row=4, column=0, pady=20, padx=20, ipadx=150, sticky="ew")

        button_5 = tk.Button(master=self.file_panel,text="Start", font=("San Francisco", 10),bg="#5D81E6",fg="#FFFFFF",command=self.startProcess)

        button_5.grid(row=5, column=0, pady=20, padx=20, ipadx=150, sticky="ew")



        self.root.mainloop()

    def openPath(self, file):
        try:        
            if file == 'ML':
                self.entry_input_ml.delete(0, 'end')
                self.entry_input_ml.insert(0, "SELECT THE ML FILE")
                self.PathML = askopenfilename()
                if self.PathML == '':
                    self.entry_input_ml.delete(0, 'end')
                    self.entry_input_ml.insert(0, "SELECT THE ML FILE")
                else:
                    self.filePath_input_ml.set(self.PathML.split('/')[-1])
                    wb = self.getSheetNames(self.PathML)
                    self.hojaPathML = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    self.hojaPathML.grid(row=0, column=1, columnspan=2, pady=10, padx=20, ipadx=50, sticky="we")
                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    #wb = self.getSheetNames(self.pathPEP)
                    #self.hojaPEP = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    #self.hojaPEP.grid(row=0, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")
            
            if file == 'REMA':
                self.entry_remaining.delete(0, 'end')
                self.entry_remaining.insert(0, "SELECT REMAINING FILE")
                self.Path_remaining = askopenfilename()
                if self.Path_remaining == '':
                    self.entry_remaining.delete(0, 'end')
                    self.entry_remaining.insert(0, "SELECT REMAINING FILE")
                else:
                    self.filePath_remaining.set(self.Path_remaining.split('/')[-1])
                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    wb = self.getSheetNames(self.Path_remaining)
                    self.hojaPathRemaining = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    self.hojaPathRemaining.grid(row=2, column=1, columnspan=2, pady=10, padx=20, ipadx=50, sticky="we")
                    
                    #wb = openpyxl.load_workbook(filename=self.pathPEP, read_only=True, keep_links=False)
                    #wb = self.getSheetNames(self.pathPEP)
                    #self.hojaPEP = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    #self.hojaPEP.grid(row=0, column=1, columnspan=1, pady=10, padx=20, ipadx=50, sticky="we")

            if file == 'restOrderBpartinfo':
                self.entry_restOrderBpartinfo.delete(0, 'end')
                self.entry_restOrderBpartinfo.insert(0, "SELECT REST ORDER BPART FILE")
                self.Path_entry_restOrderBpartinfo = askopenfilename()
                if self.Path_entry_restOrderBpartinfo == '':
                    self.entry_restOrderBpartinfo.delete(0, 'end')
                    self.entry_restOrderBpartinfo.insert(0, "SELECT REST ORDER BPART FILE")
                else:
                    self.filePath_restOrderBpartinfo.set(self.Path_entry_restOrderBpartinfo.split('/')[-1])
                    wb = self.getSheetNames(self.Path_entry_restOrderBpartinfo)
                    self.hojaPathrestOrderBpartinfo = customtkinter.CTkComboBox(self.filePanel, values=wb)
                    self.hojaPathrestOrderBpartinfo.grid(row=1, column=1, columnspan=2, pady=10, padx=20, ipadx=50, sticky="we")


        except Exception as e:
            print(e)

    def select_files_Material_list(self):
        # Abre el diálogo de selección de archivos
        filetypes = (("Archivos de Excel", "*.xls*"), ("Todos los archivos", "*.*"))
        files = filedialog.askopenfilenames(parent=self.root, title="Seleccionar archivo", filetypes=filetypes,multiple=False)
        
        if files: 
            # Imprime los archivos seleccionados como demostración
            print("Archivo seleccionado:")
            for file in files:
                self.Material_list=file
                print(self.Material_list)
                file_name = os.path.basename(file)
                print(file_name)
                self.select_files_button_Material_list.config(text=file_name)
                # Load the Excel file
                #workbook = openpyxl.load_workbook(filename=self.Material_list)
                # Create a list of sheet names
                #sheet_names = workbook.sheetnames
                # Create a dropdown menu
                #self.sheet_dropdown = tk.StringVar(self.root)
                #self.sheet_dropdown.set(sheet_names[0]) # Set the default sheet
                #self.sheet_menu = tk.OptionMenu(self.file_panel, self.sheet_dropdown, *sheet_names)
                #self.sheet_menu.grid(row=0, column=2, columnspan=2, pady=10, padx=20, ipadx=20, sticky="we")
                #tk.messagebox.showinfo("Archivo","Archivo guardado")
                return self.Material_list
        else:
            tk.messagebox.showerror("Error","Seleccione un archivo. ")

    def select_files_restOrderBpartinfo(self):
        # Abre el diálogo de selección de archivos
        filetypes = (("Archivos de Excel", "*.xls*"), ("Todos los archivos", "*.*"))
        files = filedialog.askopenfilenames(parent=self.root, title="Seleccionar archivo", filetypes=filetypes,multiple=False)
        
        if files: 
            # Imprime los archivos seleccionados como demostración
            print("Archivo seleccionado:")
            for file in files:
                self.restOrderBpartinfo=file
                print(self.restOrderBpartinfo)
                file_name = os.path.basename(file)
                print(file_name)
                self.select_files_button_restOrderBpartinfo.config(text=file_name)
                # Load the Excel file
                #workbook = openpyxl.load_workbook(filename=self.restOrderBpartinfo)
                # Create a list of sheet names
                #sheet_names = workbook.sheetnames
                # Create a dropdown menu
                #self.sheet_dropdown = tk.StringVar(self.root)
                #self.sheet_dropdown.set(sheet_names[0]) # Set the default sheet
                #self.sheet_menu1 = tk.OptionMenu(self.file_panel, self.sheet_dropdown, *sheet_names)
                #self.sheet_menu1.grid(row=1, column=2, columnspan=2, pady=10, padx=20, ipadx=20, sticky="we")
                #tk.messagebox.showinfo("Archivo","Archivo guardado")
                return self.restOrderBpartinfo
        else:
            tk.messagebox.showerror("Error","Seleccione un archivo. ")

    def select_files_Remaining(self):
        # Abre el diálogo de selección de archivos
        filetypes = (("Archivos de Excel", "*.xls*"), ("Todos los archivos", "*.*"))
        files = filedialog.askopenfilenames(parent=self.root, title="Seleccionar archivo", filetypes=filetypes,multiple=False)
        
        if files: 
            # Imprime los archivos seleccionados como demostración
            print("Archivo seleccionado:")
            for file in files:
                self.Remaining=file
                print(self.Remaining)
                file_name = os.path.basename(file)
                print(file_name)
                self.select_files_button_remaining.config(text=file_name)
                # Load the Excel file
                #workbook = openpyxl.load_workbook(filename=self.Remaining)
                # Create a list of sheet names
                #sheet_names = workbook.sheetnames
                # Create a dropdown menu
                #self.sheet_dropdown = tk.StringVar(self.root)
                #self.sheet_dropdown.set(sheet_names[0]) # Set the default sheet
                #self.sheet_menu = tk.OptionMenu(self.file_panel, self.sheet_dropdown, *sheet_names)
                #self.sheet_menu.grid(row=2, column=2, columnspan=2, pady=10, padx=20, ipadx=20, sticky="we")
                #tk.messagebox.showinfo("Archivo","Archivo guardado")
                return self.Remaining
        else:
            tk.messagebox.showerror("Error","Seleccione un archivo. ")

    def getSheetNames(self, file_path):
        sheets = []
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            xml = zip_ref.read("xl/workbook.xml").decode("utf-8")
            for s_tag in re.findall("<sheet [^>]*", xml):
                sheets.append(re.search('name="[^"]*', s_tag).group(0)[6:])
        return sheets

    def startProcess(self):
        try:                                                #card_report,sfp_report,atp_Inventario
            threading.Thread(target=get_result, args=(self.Material_list,'Sheet1',self.restOrderBpartinfo,'Sheet1',self.Remaining,'Spart Total Qty.')).start()
        except Exception as e:
            print(e)
            messagebox.showerror("Error", "An unexpected error has occurred. Please try again or contact j84319062")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    a = MainWindow()
