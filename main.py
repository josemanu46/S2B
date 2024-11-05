import sys
import os
import time
import pdb
from tkinter import messagebox
import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill
import linecache
import win32com.client
from styleframe import StyleFrame , utils, Styler

current_dir = str(os.path.dirname(os.path.realpath(__file__)))

export_path = os.path.join(current_dir, 'output')
export_file = os.path.join(export_path, f'output.xlsx')
Spart_Import_Template = os.path.join(export_path, f'Spart_Import_Template.xlsx')
Service_Product_Name_dic = os.path.join(export_path, f'Service_Product_Name_dic.txt')


def get_result(archivo_input_ml, shet_archivo_input_ml,archivo_restOrderBpartinfo, shet_archivo_restOrderBpartinfo,export_remaining_info, shet_export_remaining_info):
    print('Start Generate report...')
    # data_path = os.path.join(current_dir, 'Data')
    # Leer los archivos Excel
    #input ml
    df_input_ml = pd.read_excel(archivo_input_ml, sheet_name= shet_archivo_input_ml, usecols=['Contract','DU ID','Spart', 'QTY'],dtype={'Spart': str})
    #restorderbpart
    df_restOrderBpartinfo = pd.read_excel(archivo_restOrderBpartinfo, sheet_name=shet_archivo_restOrderBpartinfo ,usecols=['Item Code', 'Qty', 'Spart Number'], dtype={'Item Code': str, 'Spart Number': str})
    #remainig
    df_export_remaining_info = pd.read_excel(export_remaining_info, sheet_name=shet_export_remaining_info,usecols=['Huawei Contract No.','PU No.','Product Code','Product Name','Spart Number', 'Total Qty.','Margin Qty.'], dtype={'Spart Number': str,'Huawei Contract No.': str,'PU No.': str,'Product Code': str})
    
    #pdb.set_trace()
    # leer el archivo txt
    sheetParam = pd.read_csv(Service_Product_Name_dic, header=None)

    # convertir la primera columna a una lista
    sheetParam_lista = sheetParam[0].tolist()

    # Filtrar df_restOrderBpartinfo para incluir solo los valores de 'Item' de df_input_ml
    df_restOrderBpartinfo_filtered = df_restOrderBpartinfo[df_restOrderBpartinfo['Item Code'].isin(df_input_ml['Spart'])]

    #pdb.set_trace
    # Agrupar por 'Item Code' y 'Spart' y sumar las cantidades
    df_sum_qty = df_restOrderBpartinfo_filtered.groupby(['Item Code', 'Spart Number'], as_index=False)['Qty'].sum()

    # Renombrar la columna 'QTY' a 'Sum of Qty'
    df_sum_qty = df_sum_qty.rename(columns={'Qty': 'Sum of Qty'})

    # Imprimir el nuevo DataFrame
    print(df_sum_qty)

    df_export_remaining_info_sorted = df_export_remaining_info.sort_values(by='Margin Qty.', ascending=False)

    df_export_remaining_info_sorted.drop_duplicates(subset='Spart Number', keep='first', inplace=True)

    # Fusionar con df_export_remaining_info usando 'Spart Number'
    df_sum_qty = pd.merge(df_sum_qty, df_export_remaining_info_sorted, left_on='Spart Number', right_on='Spart Number', how='left')

    df_sum_qty = pd.merge(df_sum_qty, df_input_ml, left_on='Item Code', right_on='Spart', how='left')
    #pdb.set_trace()
    df_final = df_sum_qty.drop(['Spart','Contract'], axis=1)

    df_final = df_final.rename(columns={'QTY': 'QTY ML'})
    # total qty int
    df_final = df_final.fillna(0).astype({'Sum of Qty': int, 'Total Qty.': int})

    
    # Agregar una nueva columna 'Division' que es 'Sum of Qty' dividido por 'Total Qty'
    df_final['Equivalence'] = df_final['Sum of Qty'] / df_final['Total Qty.']
    # Reemplaza los valores infinitos (positivos y negativos) en la columna 'Equivalence' por NaN (valor faltante):
    df_final['Equivalence'] = df_final['Equivalence'].replace([np.inf, -np.inf], np.nan)
    # Rellena los valores faltantes en el DataFrame con 0:
    df_final = df_final.fillna(0)

    # Redondea los valores en la columna 'Equivalence' a la unidad más cercana:
    df_final['Equivalence'] = df_final['Equivalence'].round(1)

    # Crea una nueva columna 'S2B' en el DataFrame df_final y calcula su valor como el cociente de la columna 'QTY ML' dividido entre la columna
    df_final['S2B'] = df_final['QTY ML'] / df_final['Equivalence']

    # Redondea los valores en la columna 'S2B' a la unidad más cercana:
    df_final['S2B'] = df_final['S2B'].round(1)
    # Reemplaza los valores 0 en la columna 'S2B' por 1:
    #df_final['S2B'] = df_final['S2B'].replace(0, 1)

    # Convertir NaN a 0 y las columnas float a int Convierte los valores faltantes en el DataFrame a 0 y convierte las columnas de tipo float a int:
    #df_final = df_final.fillna(0).astype({'S2B': int})

    #Crea una nueva columna 'Status' en el DataFrame df_final y asigna el valor 'ok' si el valor en la columna 'Margin Qty.' es mayor o igual al valor en la columna 'S2B', y 'error'   en    caso contrario:
    # Convierte la columna 'Margin Qty.' en un tipo de datos numérico
    #df_final['Margin Qty.'] = pd.to_numeric(df_final['Margin Qty.'], errors='coerce')


    # Suponiendo que df_final es tu DataFrame y 'Spart Number' y 'Item Code' son las columnas que quieres verificar
    #duplicados = df_final.loc[(df_final['Spart Number'].duplicated()) & (df_final['Item Code'] == df_final['Spart Number']), ['Spart Number', 'Item Code']]



    # Agrega una nueva condición para verificar si el valor de 'S2B' es decimal
    df_final['Status'] = np.where(df_final['Margin Qty.'].notnull() & df_final['S2B'].apply(np.isreal), np.where(df_final['Margin Qty.'] >= df_final['S2B'], 'ok', 'No Margin'), 'No Margin')
    
    # Crea una nueva columna llamada 'status'
    df_final['Status'] = np.where(df_final['Item Code'] == df_final['Spart Number'], 'ok', 'aux')

    #df_final['Status'] = np.where(df_final['Margin Qty.'] >= df_final['S2B'], 'ok', 'error')

    # crear la nueva columna 'tipo de servicio'
    df_final['Product Type'] = df_final['Product Name'].apply(lambda x: 'Service Material' if x in sheetParam_lista else 'Equipment')

    
    df_final = df_final.assign(Spart_Number_copy = df_final['Spart Number'].copy())
    df_final = df_final.rename(columns={'Spart_Number_copy': 'Spart_Number'})
    # Encontrar los índices de los elementos que son duplicados
    indices = list(df_final['Spart_Number'][df_final['Spart_Number'].duplicated(keep=False)].index)

    print(indices)
    columns = ['Item Code','Spart Number', 'Sum of Qty','Total Qty.','QTY ML','Equivalence','Huawei Contract No.','PU No.','DU ID','Product Code','Product Name','Product Type','Spart_Number','S2B','Margin Qty.','Status']

    col_blue = ['Item Code','Spart Number', 'Sum of Qty','Total Qty.','QTY ML']
    col_red = ['Equivalence']
    col_green = ['Huawei Contract No.','PU No.','DU ID','Product Code','Product Name','Spart_Number','Product Type','S2B']
    col_orange = ['Margin Qty.','Status']
    df_final = df_final[columns]

    # Guardar el nuevo DataFrame en un archivo Excel y sobrescribirlo
    df_final.to_excel(export_file, index=False)
    
    df_frame_full = pd.read_excel(export_file, sheet_name= 0, header=0)
    #df_frame_full[''] = df_frame_full['PU No.'].astype(str)
    df_frame_full['PU No.'] = df_frame_full['PU No.'].apply(lambda x: "00" + str(x))

    excel_writer = StyleFrame.ExcelWriter(export_file)
    sf = StyleFrame(df_frame_full)
    
    sf.apply_headers_style(cols_to_style = col_blue,styler_obj=Styler(bg_color='#95B9D7', font_color='black'))
    sf.apply_headers_style(cols_to_style = 'Equivalence' ,styler_obj=Styler(bg_color='#FFC7CE', font_color='black'))
    sf.apply_headers_style(cols_to_style = col_green,styler_obj=Styler(bg_color='#C6EFCE', font_color='black'))
    sf.apply_headers_style(cols_to_style = col_orange,styler_obj=Styler(bg_color='#FFEB9C', font_color='black'))
    sf.apply_style_by_indexes(indexes_to_style=indices,cols_to_style='Spart_Number', styler_obj=Styler(bg_color='#FFC7CE', font_color='black'))

    # Colorear celdas con valores duplicados en la columna 'spart number'
    sf.to_excel(excel_writer=excel_writer,best_fit=columns)
     
    excel_writer.save()
    # Imprimir la ruta del archivo de salida
    print(f"Archivo guardado en: {export_file}")

    df_filtrado_template = df_frame_full[df_frame_full['Status'] == 'ok']

    #template_file = f'nombre_del_template_{proyecto}.xlsx'  # Reemplaza con el nombre real de tus templates
    fill_teamplate(Spart_Import_Template, df_filtrado_template)
    
    messagebox.showinfo("Information", "File created successfully")
    abrir_ubicacion(export_path)

def abrir_ubicacion(template):
    if template: 
        respuesta = messagebox.askyesno("Confirm", "Do you want to open the folder?")
        if respuesta:
            try:
                os.startfile(template)
            except OSError as e:
                messagebox.showerror("Error", f"Could not open the folder location:\n{str(e)}")
    else:
        messagebox.showinfo("Information", "A folder location has not been selected.")

def fill_teamplate(templatefile,result_df):
    export_template = os.path.join(export_path,templatefile)
    template = openpyxl.load_workbook(filename=export_template, data_only=True)
    template_sheet = template['Spart Configuration']
    for row in template_sheet.iter_rows(min_row=2, max_row=template_sheet.max_row, max_col=template_sheet.max_column):
        for cell in row:
            cell.value = None
    mapeo_columnas = {
         'Huawei Contract No.': 'A',
         'Product Type': 'B',
         'PU No.':'E',
         'DU ID': 'F',
         'Product Code': 'H',
         'Product Name': 'I',
         'Spart_Number': 'J',
         'S2B': 'K'
     }
    for col_result, col_template in mapeo_columnas.items():
        for i, valor in enumerate(result_df[col_result], start=2):
            #colorear los valores duplicados sobrte el spart number
            # if col_result == 'Spart_Number' and result_df.duplicated(subset='Spart_Number', keep=False).iloc[i-2]:
            #     template_sheet[f'{col_template}{i}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            template_sheet[f'{col_template}{i}'] = valor

    template.save(os.path.join(export_path, f'Spart_Import_Template.xlsx'))

